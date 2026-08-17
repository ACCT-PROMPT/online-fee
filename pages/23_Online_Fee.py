"""
pages/23_Online_Fee.py
รวม tool ดึงค่าธรรมเนียม Platform ทุกตัวไว้ในที่เดียว
Tabs: KBank · KTC · Shopee(รวม) · Shopee(รายชิ้น) · SPX · TikTok(นายหน้า) · TikTok(ค่าธรรมเนียม) · TikTok(ขนส่ง)
"""
import io, os, re, sys, tempfile
from pathlib import Path

import streamlit as st

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import style

# เพิ่ม path ของ run\online scripts
_ONLINE_SRC = r"C:\Users\User\Desktop\run\online"
if _ONLINE_SRC not in sys.path:
    sys.path.insert(0, _ONLINE_SRC)

st.set_page_config(page_title="Online Fee", page_icon="🌐", layout="wide")
style.inject()
style.back_home()

st.markdown(
    '<span style="background:linear-gradient(135deg,#1a3a6b,#2c5aa0);color:white;'
    'border-radius:20px;padding:4px 14px;font-size:12px;font-weight:700;">PDF → Excel</span>',
    unsafe_allow_html=True,
)
st.title("🌐 Online Fee Extractor")
st.caption("ดึงข้อมูลค่าธรรมเนียม Platform จาก PDF → Excel — ครบทุก Platform ในที่เดียว")


# ─── shared helper ─────────────────────────────────────────────────────────────

def _uploader(key: str, label: str = "เลือกไฟล์ PDF (หลายไฟล์ได้)"):
    """File uploader + clear button — คืน list[UploadedFile] หรือ []
    key ต้องไม่ซ้ำกันระหว่าง tab เพราะ Streamlit render ทุก tab พร้อมกัน
    """
    _uk = style.upload_key(key)
    _, _cb = st.columns([5, 1])
    with _cb:
        style.clear_files_button(key)
    files = st.file_uploader(
        label, type="pdf", accept_multiple_files=True,
        key=f"files_{key}_{_uk}", label_visibility="collapsed",
    )
    return files or []


def _run_online_script(script_name: str, uploaded_files) -> bytes | None:
    """
    บันทึกไฟล์ที่อัปโหลดลง temp dir → เรียก process_selected_pdfs() → คืน bytes ของ Excel
    """
    import importlib
    mod = importlib.import_module(script_name)
    importlib.reload(mod)          # reload เพื่อหลีกเลี่ยง state ค้าง

    with tempfile.TemporaryDirectory() as tmpdir:
        file_paths = []
        for f in uploaded_files:
            dst = os.path.join(tmpdir, f.name)
            with open(dst, "wb") as fp:
                fp.write(f.read())
            file_paths.append(dst)

        out_path = mod.process_selected_pdfs(file_paths)
        if not out_path or not Path(out_path).exists():
            return None
        return Path(out_path).read_bytes()


def _download_btn(excel_bytes: bytes, filename: str, n: int, n_files: int):
    st.success(f"สำเร็จ — พบข้อมูล {n} รายการจาก {n_files} ไฟล์")
    st.download_button(
        "📥 ดาวน์โหลด Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ─── tabs ──────────────────────────────────────────────────────────────────────

tabs = st.tabs([
    "🏦 KBank",
    "💳 KTC",
    "🛒 Shopee (รวม)",
    "🛍️ Shopee (รายชิ้น)",
    "🚚 SPX",
    "🎵 TikTok (นายหน้า)",
    "📦 TikTok (ค่าธรรมเนียม)",
    "🚛 TikTok (ขนส่ง)",
])


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — KBank Fee
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[0]:
    st.subheader("🏦 KBank Fee Extractor")
    st.caption("ดึงข้อมูลค่าธรรมเนียม KBank จาก PDF → Excel")

    import fitz, openpyxl
    from openpyxl.styles import Font as OFont

    def _kbank_extract(pdf_bytes: bytes, filename: str) -> list[dict]:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        rows = []
        for page_num, page in enumerate(doc, start=1):
            text = page.get_text("text")
            data = {"File Name": filename, "Page No.": page_num,
                    "Issued Date": "", "Document Number": "", "Payment Type": "",
                    "Fee": None, "VAT": None, "Total (Fee+VAT)": None}
            m = re.search(r"([A-Z0-9]{8,})", text)
            if m: data["Document Number"] = m.group(1)
            m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
            if m: data["Issued Date"] = m.group(1)
            m = re.search(r"(บัตรเครดิต/เดบิต|กระเป๋าเงินอิเล็กทรอนิกส์)", text)
            if m: data["Payment Type"] = m.group(1)
            m = re.search(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})", text)
            if m:
                _, fee, vat, _ = m.groups()
                data["Fee"] = float(fee.replace(",", ""))
                data["VAT"] = float(vat.replace(",", ""))
                data["Total (Fee+VAT)"] = data["Fee"] + data["VAT"]
            rows.append(data)
        return rows

    def _kbank_excel(data: list[dict]) -> bytes:
        headers = ["File Name", "Page No.", "Issued Date", "Document Number",
                   "Payment Type", "Fee", "VAT", "Total (Fee+VAT)"]
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PDF Data"
        ws.append(headers)
        for row in data:
            ws.append([row[h] for h in headers])
        for col in ["F", "G", "H"]:
            for cell in ws[col][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
        last = ws.max_row + 1
        ws[f"E{last}"] = "Grand Total"; ws[f"E{last}"].font = OFont(bold=True)
        for col in ["F", "G", "H"]:
            ws[f"{col}{last}"] = f"=SUM({col}2:{col}{last-1})"
            ws[f"{col}{last}"].number_format = "#,##0.00"
            ws[f"{col}{last}"].font = OFont(bold=True)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    files_kb = _uploader("kbank_fee")
    if files_kb and st.button("Extract → Excel", key="btn_kbank", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            all_data = []
            for f in files_kb:
                all_data.extend(_kbank_extract(f.read(), f.name))
        if all_data:
            _download_btn(_kbank_excel(all_data), "KBank_Fee_Extracted.xlsx",
                          len(all_data), len(files_kb))
        else:
            st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — KTC Fee
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[1]:
    st.subheader("💳 KTC Fee Extractor")
    st.caption("ดึงข้อมูลค่าธรรมเนียม KTC จาก PDF → Excel")

    import pdfplumber
    import pandas as pd

    TH_COLS_KTC = ["ชื่อไฟล์", "เลขหน้า", "วันที่", "เลขที่", "รายการ",
                   "รวม / TOTAL", "ภาษีมูลค่าเพิ่ม 7% / VAT 7%", "จำนวนเงินรวม / GRAND TOTAL"]
    _AMT_RE   = re.compile(r"([\d,]+(?:\.\d{2})?)")
    _TOTAL_LB = re.compile(r"รวม\s*/\s*TOTAL\b", re.IGNORECASE)
    _VAT_LB   = re.compile(r"(ภาษีมูลค่าเพิ่ม\s*7%\s*/\s*VAT\s*7%|VAT\s*7%)", re.IGNORECASE)
    _GRAND_LB = re.compile(r"(จำนวนเงินรวม\s*/\s*GRAND\s*TOTAL|GRAND\s*TOTAL)", re.IGNORECASE)
    _DESC_BL  = re.compile(r"(SETTLEMENT\s*DATE|TRANSACTION\s*DATE|วันที่สรุปยอด|วันที่ทำรายการ|"
                            r"รวม\s*/\s*TOTAL|VAT|ภาษีมูลค่าเพิ่ม|จำนวนเงินรวม|GRAND\s*TOTAL)", re.IGNORECASE)

    def _ktc_norm(s): return re.sub(r"[ \t]+", " ", s or "").strip()
    def _ktc_last_amt(line):
        nums = _AMT_RE.findall(line or "")
        return nums[-1].replace(",", "") if nums else None
    def _ktc_to_num(s):
        if s is None: return ""
        try: return float(str(s).strip().replace(",", ""))
        except: return ""

    def _ktc_parse(pdf_bytes: bytes, filename: str) -> list[dict]:
        rows = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                lines = [_ktc_norm(l) for l in text.replace("\r", "\n").split("\n") if l.strip()]
                # issue date
                m = re.search(r"วันที่\s*/\s*ISSUE\s*DATE\s*:\s*([^\n\r]+)", text, re.IGNORECASE)
                date = _ktc_norm(m.group(1)) if m else ""
                # doc no
                m = re.search(r"เลขที่\s*/\s*NO\.?\s*:\s*([^\n\r]+)", text, re.IGNORECASE)
                doc = _ktc_norm(m.group(1)) if m else ""
                # first item description
                desc = ""
                for line in lines:
                    mm = re.match(r"^1\s+(.+?)\s+[\d,]+(?:\.\d{2})?\s*$", line)
                    if mm:
                        d = _ktc_norm(mm.group(1))
                        if d and not _DESC_BL.search(d):
                            desc = d; break
                if not desc: continue
                # totals
                total = vat = grand = None
                for idx, line in enumerate(lines):
                    if total is None and _TOTAL_LB.search(line) and not _GRAND_LB.search(line):
                        total = _ktc_last_amt(line) or (_ktc_last_amt(lines[idx+1]) if idx+1<len(lines) else None)
                    if vat is None and _VAT_LB.search(line):
                        vat = _ktc_last_amt(line) or (_ktc_last_amt(lines[idx+1]) if idx+1<len(lines) else None)
                    if grand is None and _GRAND_LB.search(line):
                        grand = _ktc_last_amt(line) or (_ktc_last_amt(lines[idx+1]) if idx+1<len(lines) else None)
                rows.append({"ชื่อไฟล์": filename, "เลขหน้า": i, "วันที่": date, "เลขที่": doc,
                             "รายการ": desc, "รวม / TOTAL": _ktc_to_num(total),
                             "ภาษีมูลค่าเพิ่ม 7% / VAT 7%": _ktc_to_num(vat),
                             "จำนวนเงินรวม / GRAND TOTAL": _ktc_to_num(grand)})
        return rows

    def _ktc_excel(all_rows: list[dict]) -> bytes:
        df = pd.DataFrame(all_rows, columns=TH_COLS_KTC)
        summary = {"ชื่อไฟล์": "", "เลขหน้า": "", "วันที่": "", "เลขที่": "", "รายการ": "รวมทั้งหมด",
                   "รวม / TOTAL": df["รวม / TOTAL"].apply(lambda x: float(x) if isinstance(x, (int,float)) else 0).sum(),
                   "ภาษีมูลค่าเพิ่ม 7% / VAT 7%": df["ภาษีมูลค่าเพิ่ม 7% / VAT 7%"].apply(lambda x: float(x) if isinstance(x,(int,float)) else 0).sum(),
                   "จำนวนเงินรวม / GRAND TOTAL": df["จำนวนเงินรวม / GRAND TOTAL"].apply(lambda x: float(x) if isinstance(x,(int,float)) else 0).sum()}
        if not df.empty:
            df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)
            sr = len(df) + 1
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Extracted")
            ws = writer.sheets["Extracted"]
            for cl, cn in zip("ABCDEFGH", TH_COLS_KTC):
                ws.column_dimensions[cl].width = min(max(12, max((len(str(v)) for v in df[cn].astype(str)), default=10) + 2), 60)
            from openpyxl.styles import Font as OFont2
            for ci in [6, 7, 8]:
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.00"
            if not df.empty:
                for ci2 in range(1, len(TH_COLS_KTC) + 1):
                    ws.cell(row=sr, column=ci2).font = OFont2(bold=True)
        return buf.getvalue()

    files_ktc = _uploader("ktc_fee")
    if files_ktc and st.button("Extract → Excel", key="btn_ktc", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            all_rows = []
            for f in files_ktc:
                all_rows.extend(_ktc_parse(f.read(), f.name))
        if all_rows:
            _download_btn(_ktc_excel(all_rows), "KTC_Fee_Extracted.xlsx",
                          len(all_rows), len(files_ktc))
        else:
            st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Shopee Fee (ยอดรวม)
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[2]:
    st.subheader("🛒 Shopee Fee Extractor (ยอดรวม)")
    st.caption("ดึงรายการค่าธรรมเนียม Shopee ยอดรวมต่อเอกสาร จาก PDF → Excel")

    from openpyxl import Workbook as OWB
    from openpyxl.styles import Font as OFont3

    def _shopee_fee_extract(pdf_bytes: bytes, filename: str) -> list[dict]:
        from datetime import datetime as DT
        entries = []
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        except Exception as e:
            st.warning(f"เปิดไฟล์ไม่ได้: {filename} — {e}"); return entries
        for page_num, page in enumerate(doc, start=1):
            text = page.get_text()
            doc_no = ""
            m = re.search(r'เลขที่/ No\.\s*([A-Z0-9\-]+)[\n\s]*([0-9\-]+)', text)
            if m: doc_no = (m.group(1) + m.group(2)).replace(" ", "").replace("\n", "")
            date_str = ""
            m = re.search(r'วันที่/ Date\s*(\d{2}/\d{2}/\d{4})', text)
            if m:
                try: date_str = DT.strptime(m.group(1), "%d/%m/%Y").strftime("%Y-%m-%d")
                except: date_str = m.group(1)
            item_matches = re.findall(r'\d+\s+([^\n]+?)\s+\d+\s+[\d,]+\.\d+\s+([\d,]+\.\d+)', text)
            paid_ads_amount = None; page_entries = []
            for item_name, amount in item_matches:
                item_name = item_name.strip()
                amount_value = float(amount.replace(",", ""))
                if item_name.lower() == "paid ads":
                    paid_ads_amount = amount_value
                else:
                    page_entries.append({"ไฟล์": filename, "เลขหน้า": page_num,
                                         "เลขที่เอกสาร": doc_no, "วันที่": date_str,
                                         "รายการ": item_name, "จำนวนเงิน": amount_value})
            m_disc = re.search(r'(สวนลด|ส่วนลด|Discount)\s*[:\-]?\s*([\d,]+\.\d+)', text)
            discount = float(m_disc.group(2).replace(",", "")) if m_disc else 0.0
            if paid_ads_amount is not None:
                page_entries.append({"ไฟล์": filename, "เลขหน้า": page_num,
                                     "เลขที่เอกสาร": doc_no, "วันที่": date_str,
                                     "รายการ": "Paid ads", "จำนวนเงิน": paid_ads_amount - discount})
            entries.extend(page_entries)
            m_total = re.search(r'Total Value of Services \(Included VAT\)\s*([\d,]+\.\d+)', text)
            if m_total:
                entries.append({"ไฟล์": filename, "เลขหน้า": page_num,
                                 "เลขที่เอกสาร": doc_no, "วันที่": date_str,
                                 "รายการ": "Total Value of Services (Included VAT)",
                                 "จำนวนเงิน": float(m_total.group(1).replace(",", ""))})
        return entries

    def _shopee_fee_excel(all_entries: list[dict]) -> bytes:
        df = pd.DataFrame(all_entries)
        wb = OWB(); ws = wb.active; ws.title = "Shopee Items"
        ws.append(list(df.columns))
        for ri, row in enumerate(df.itertuples(index=False), start=2):
            for ci, value in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=value)
                if ws.cell(row=1, column=ci).value == "จำนวนเงิน":
                    cell.number_format = "#,##0.00"
                if row.รายการ == "Total Value of Services (Included VAT)":
                    cell.font = OFont3(bold=True)
        sr = ws.max_row + 1
        ws.cell(row=sr, column=5, value="รวมทั้งหมด:")
        sc = ws.cell(row=sr, column=6, value=f"=SUBTOTAL(9,F2:F{sr-1})")
        sc.font = OFont3(bold=True); sc.number_format = "#,##0.00"
        for col in ws.columns:
            mw = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(mw + 2, 50)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    files_sf = _uploader("shopee_fee_tab")
    if files_sf and st.button("Extract → Excel", key="btn_shopee_fee", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            all_entries = []
            for f in files_sf:
                all_entries.extend(_shopee_fee_extract(f.read(), f.name))
        if all_entries:
            _download_btn(_shopee_fee_excel(all_entries), "Shopee_Fee_Extracted.xlsx",
                          len(all_entries), len(files_sf))
        else:
            st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — Shopee Items (รายชิ้น) — via run\online script
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[3]:
    st.subheader("🛍️ Shopee Items (รายชิ้น)")
    st.caption("ดึงค่าธรรมเนียม Shopee แบบรายชิ้นพร้อม Subtotal จาก PDF → Excel")

    files_si = _uploader("shopee_items_tab")
    if files_si and st.button("Extract → Excel", key="btn_shopee_items", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            try:
                excel_bytes = _run_online_script(
                    "shopee_extract_items_allpdf_with_subtotal", files_si
                )
                if excel_bytes:
                    _download_btn(excel_bytes, "shopee_all_items_with_subtotal.xlsx",
                                  "—", len(files_si))
                else:
                    st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาด: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 — SPX Shipping — via run\online script
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[4]:
    st.subheader("🚚 ค่าขนส่ง SPX")
    st.caption("ดึงข้อมูลค่าขนส่ง SPX จาก PDF → Excel")

    files_spx = _uploader("spx_tab")
    if files_spx and st.button("Extract → Excel", key="btn_spx", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            try:
                excel_bytes = _run_online_script("spx_extract_shipping_allpdf", files_spx)
                if excel_bytes:
                    _download_btn(excel_bytes, "spx_shipping_extracted.xlsx",
                                  "—", len(files_spx))
                else:
                    st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาด: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6 — TikTok Commission / Pivot — inline (same logic as 9_Commission_Pivot)
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[5]:
    st.subheader("🎵 ค่านายหน้า TikTok + Pivot")
    st.caption("ดึงค่านายหน้า TikTok พร้อม Pivot ต่อ Tax Number จาก PDF → Excel")

    from openpyxl import Workbook as OWB2
    from openpyxl.styles import Font as OFont4
    from openpyxl.utils.dataframe import dataframe_to_rows as df2rows
    from openpyxl.utils import get_column_letter as gcl

    def _tiktok_comm_extract(pdf_bytes: bytes, filename: str) -> list[tuple]:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        results = []
        for page in doc:
            text = page.get_text()
            receipt_blocks = re.findall(
                r"Receipt.*?Total Amount\s+฿[\d,]+\.\d{2}", text, flags=re.DOTALL)
            for receipt in receipt_blocks:
                def _get(pattern, default="N/A"):
                    mm = re.search(pattern, receipt, re.DOTALL | re.IGNORECASE)
                    return mm.group(1).strip() if mm else default
                period        = _get(r"Period\s*:\s*([A-Za-z]+\s\d{1,2},\s\d{4}\s-\s[A-Za-z]+\s\d{1,2},\s\d{4})")
                receipt_date  = _get(r"Receipt Date\s*:\s*([A-Za-z]{3,9} \d{1,2}, \d{4})")
                client        = _get(r"Client Name:\s*(.+)")
                username      = _get(r"Creator Username:\s*(\S+)")
                addr_m        = re.search(r"Billing Address:\s*(.+?)\s*(?:Bill To|Tax Number:)", receipt, re.DOTALL)
                address       = re.sub(r"\s+", " ", addr_m.group(1).strip()) if addr_m else "N/A"
                receipt_no    = _get(r"Receipt Number\s*:\s*(\S+)")
                tax_m         = re.search(r"(?i)Bill From.*?Tax Number:\s*(\d{13})", receipt, re.DOTALL)
                tax_number    = tax_m.group(1).strip() if tax_m else "N/A"
                for comm_type, amt_str in re.findall(
                    r"([A-Za-z ]+commission)\s*/\s*฿([\d,]+\.\d{2})", receipt, re.IGNORECASE
                ):
                    results.append((filename, receipt_no, receipt_date, period,
                                    client, username, address, tax_number,
                                    comm_type.strip().title(), float(amt_str.replace(",", ""))))
        return results

    def _tiktok_comm_excel(all_data: list[tuple]) -> bytes:
        columns = ["PDF File", "Receipt Number", "Receipt Date", "Period",
                   "Client Name", "Creator Username", "Billing Address",
                   "Tax Number", "Commission Type", "Amount (THB)"]
        df_all = pd.DataFrame(all_data, columns=columns)
        wb2 = OWB2(); ws_main = wb2.active; ws_main.title = "All Commissions"
        for r in df2rows(df_all, index=False, header=True):
            ws_main.append(r)
        last_r = ws_main.max_row
        ws_main.cell(row=last_r+1, column=1, value="TOTAL (SUBTOTAL)")
        tc = ws_main.cell(row=last_r+1, column=10, value=f"=SUBTOTAL(9,J2:J{last_r})")
        tc.font = OFont4(bold=True)
        for row in ws_main.iter_rows(min_row=2, min_col=10, max_col=10):
            for cell in row: cell.number_format = "#,##0.00"
        ws_main.freeze_panes = "B2"; ws_main.auto_filter.ref = ws_main.dimensions
        if not df_all.empty:
            pivot_df = df_all.pivot_table(
                index=["Tax Number", "Client Name", "Billing Address"],
                columns="Commission Type", values="Amount (THB)",
                aggfunc="sum", fill_value=0).reset_index()
            pivot_df.columns.name = None
            ws_p = wb2.create_sheet(title="Pivot Wide")
            for r in df2rows(pivot_df, index=False, header=True):
                ws_p.append(r)
            tot_row = ws_p.max_row + 1
            for cn in range(4, ws_p.max_column + 1):
                cl = gcl(cn)
                c = ws_p.cell(row=tot_row, column=cn, value=f"=SUBTOTAL(9,{cl}2:{cl}{tot_row-1})")
                c.font = OFont4(bold=True)
            ws_p.cell(row=tot_row, column=1, value="TOTAL (SUBTOTAL)")
            ws_p.freeze_panes = "B2"; ws_p.auto_filter.ref = ws_p.dimensions
        for ws in wb2.worksheets:
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 17.5
        buf = io.BytesIO(); wb2.save(buf); return buf.getvalue()

    files_tc = _uploader("tiktok_comm")
    if files_tc and st.button("Extract → Excel", key="btn_tiktok_comm", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            all_data = []
            for f in files_tc:
                all_data.extend(_tiktok_comm_extract(f.read(), f.name))
        if all_data:
            _download_btn(_tiktok_comm_excel(all_data), "commission_pivot_wide.xlsx",
                          len(all_data), len(files_tc))
        else:
            st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 7 — TikTok ค่าธรรมเนียม — via run\online script
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[6]:
    st.subheader("📦 ค่าธรรมเนียม TikTok")
    st.caption("ดึงข้อมูลค่าธรรมเนียม TikTok จาก PDF → Excel")

    files_tf = _uploader("tiktok_fee_tab")
    if files_tf and st.button("Extract → Excel", key="btn_tiktok_fee", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            try:
                excel_bytes = _run_online_script("tiktok_extract_fees_allpdf", files_tf)
                if excel_bytes:
                    _download_btn(excel_bytes, "tiktok_fees_extracted.xlsx",
                                  "—", len(files_tf))
                else:
                    st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาด: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 8 — TikTok ค่าขนส่ง — via run\online script
# ═══════════════════════════════════════════════════════════════════════════════
with tabs[7]:
    st.subheader("🚛 ค่าขนส่ง TikTok")
    st.caption("ดึงข้อมูลค่าขนส่ง TikTok จาก PDF → Excel")

    files_ts = _uploader("tiktok_ship_tab")
    if files_ts and st.button("Extract → Excel", key="btn_tiktok_ship", type="primary"):
        with st.spinner("กำลังประมวลผล..."):
            try:
                excel_bytes = _run_online_script("tiktok_extract_shipping_allpdf", files_ts)
                if excel_bytes:
                    _download_btn(excel_bytes, "tiktok_shipping_extracted.xlsx",
                                  "—", len(files_ts))
                else:
                    st.warning("ไม่พบข้อมูลในไฟล์ PDF ที่เลือก")
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาด: {e}")
